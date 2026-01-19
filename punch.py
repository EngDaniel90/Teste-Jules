import os
import sys
import time
import shutil
import urllib3
import urllib.parse
import requests
import pandas as pd
from datetime import datetime
import re

# openpyxl imports
import openpyxl
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter

# Importação para comunicação com Outlook Local
try:
    import win32com.client as win32
except ImportError:
    print("ERRO: Instale a biblioteca pywin32 executando: pip install pywin32")

# Selenium imports
from selenium import webdriver
from selenium.webdriver.edge.service import Service as EdgeService
from selenium.webdriver.edge.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

# --- CONFIGURAÇÕES DE REDE CORPORATIVA ---
os.environ['WDM_SSL_VERIFY'] = '0'
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

# --- CONFIGURAÇÕES DO AMBIENTE ---
URL_LOGIN_SEATRIUM = "https://seatrium.sharepoint.com/sites/P84P85DesignReview/Lists/DR90%20EHouse%20Punchlist/AllItems.aspx"
URL_BASE_SHAREPOINT = "https://seatrium.sharepoint.com/sites/P84P85DesignReview"

# --- CONFIGURAÇÃO DAS PASTAS DE DESTINO (LISTA) ---
PASTAS_DESTINO = [
    r"C:\Users\E797\PETROBRAS\SRGE SI-II SCP85 ES - Planilha_BI_Punches",
    r"C:\Users\E797\Downloads\Teste mensagem e print"
]

CAMINHO_DRIVER_FIXO = r"C:\Users\E797\PycharmProjects\pythonProject\msedgedriver.exe"
EMAIL_DESTINO = '658b4ef7.petrobras.com.br@br.teams.ms'

# --- CONFIGURAÇÃO DAS LISTAS SHAREPOINT ---
LISTAS_SHAREPOINT = {
    "Topside": {
        "nome_api": "P84/85_TOPSIDE_DR90_Punch_List",
        "arquivo_saida": "Punch_DR90_TS.xlsx",
        "colunas": [
            "DECK No.", "Action Description", "KBR Comment", "Company", "KBR Discipline", "Status",
            "Date Cleared by KBR", "Petrobras Response By", "Petrobras Response Date", "Petrobras Response ",
            "Petrobras Remarks", "Petrobras Discipline", "Petrobras Responsible", "Seatrium Remarks", "Zone",
            "Date Cleared by Petrobras", "S3D Item Tags", "Punch No", "KBR Target Date",
            "Days Since Date Cleared by KBR",
            "Days Since Date Cleared by Seatrium", "Punched by (Group)", "Petrobras Need Operation to close? (Y/N)",
            "Date Cleared by Petrobras Operation", "Petrobras Operation accept closing? (Y/N)", "Is Reopen? (Y/N)",
            "Seatrium Target Date Calculated", "Petrobras Operation Target Date Calculated",
            "Petrobras Target Date Calculated", "Petrobras Target Date", "Petrobras Operation Target Date",
            "Seatrium Target Date"
        ]
    },
    "E-House": {
        "nome_api": "DR90 E-House Punchlist",
        "arquivo_saida": "Punch_DR90_E-House.xlsx",
        "colunas": [
            "Punch No", "Zone", "DECK No.", "Zone-Punch Number", "Action Description", "Punched by", "Punch SnapShot1",
            "Punch SnapShot2", "Closing SnapShot1", "Hotwork", "ABB/CIMC Discipline", "Company", "Close Out Plan Date",
            "Action by", "Status", "Action Comment", "Date Cleared by ABB", "Days Since Date Cleared by ABB",
            "KBR Response", "KBR Response Date", "KBR Response by", "KBR Remarks", "KBR Category", "KBR Discipline",
            "KBR Screenshot", "Date Cleared by KBR", "Days Since Date Cleared By KBR", "Seatrium Discipline",
            "Seatrium Remarks", "Checked By (Seatrium)", "Seatrium Comments",
            ("DateClearBySeatrium", "Date Cleared By Seatrium"),
            "Days Since Date Cleared by Seatrium", "Petrobras Response", "Petrobras Response By",
            "Petrobras Screenshot",
            "Petrobras Response Date", "Petrobras Remarks", "Petrobras Discipline", "Petrobras Category",
            "Date Cleared by Petrobras", "Days Since Date Cleared By Petrobras", "Additional Remarks",
            "ARC Reference No(HFE Only)", "Modified", "Modified By", "Item Type", "Path"
        ]
    },
    "Vendors": {
        "nome_api": "Vendor Package Review Punchlist DR90",
        "arquivo_saida": "Punch_DR90_Vendors.xlsx",
        "colunas": [
            "Punch No", "Zone", "DECK No.", "Zone-Punch Number", "Action Description", "S3D Item Tags", "Punched by",
            "Punch Snapshot", "Punch Snapshot 2", "Punch Snapshot 3", "Punch Snapshot 4", "Close-Out Snapshot 1",
            "Close-Out Snapshot 2", "Action Comment", "Vendor Discipline", "Company", "Action by", "Status",
            "Date Cleared by KBR", "Days Since Date Cleared by KBR", "Petrobras Response", "Petrobras Response by",
            "Petrobras Response Date", "Petrobras Screenshot", "Remarks", "Petrobras Discipline", "Petrobras Category",
            "Date Cleared by Petrobras", "Seatrium Remarks", "Seatrium Discipline", "Checked By (Seatrium)",
            "Seatrium Comments", ("DateClearBySeatrium", "Date Cleared By Seatrium"),
            "Days Since Date Cleared by Seatrium", "Modified By",
            "Item Type", "Path"
        ]
    }
}


class AutomacaoPunchList:
    def __init__(self):
        self.driver = None
        self.log_sessao = []
        self.schema_lista = {}

    def registrar_log(self, mensagem):
        timestamp = datetime.now().strftime('%H:%M:%S')
        texto = f"[{timestamp}] {mensagem}"
        print(texto)
        self.log_sessao.append(texto)

    def normalize_key(self, text):
        """Remove símbolos e espaços para comparação robusta."""
        if not text:
            return ""
        # Mantém apenas letras e números, minúsculo
        return "".join(c for c in text if c.isalnum()).lower()

    def enviar_via_outlook_app(self, sucesso):
        status_geral = "SUCESSO" if sucesso else "FALHA"
        cor_status_geral = "#28a745" if sucesso else "#dc3545"

        log_html_lines = []
        for linha in self.log_sessao:
            classe_css = ""
            linha_sem_ts = linha.split("] ", 1)[-1] if "] " in linha else linha

            if "SUCESSO" in linha_sem_ts:
                classe_css = "success"
            elif "ERRO" in linha_sem_ts or "FALHA" in linha_sem_ts or "Falha" in linha_sem_ts:
                classe_css = "error"
            elif "AVISO" in linha_sem_ts:
                classe_css = "warning"
            elif "---" in linha_sem_ts:
                classe_css = "info"

            log_html_lines.append(f'<div class="log-line {classe_css}">{linha}</div>')

        log_html = "".join(log_html_lines)

        html_body = f"""
        <html>
        <head>
            <style>
                body {{ font-family: Arial, sans-serif; color: #333; }}
                .container {{ padding: 20px; border: 1px solid #dee2e6; border-radius: 5px; max-width: 900px; margin: auto; }}
                .header {{ font-size: 24px; font-weight: bold; color: #004085; border-bottom: 2px solid #004085; padding-bottom: 10px; margin-bottom: 20px;}}
                .status {{ font-size: 20px; font-weight: bold; padding: 12px; color: white; background-color: {cor_status_geral}; border-radius: 4px; text-align: center; }}
                .log-container {{ margin-top: 25px; padding: 20px; background-color: #f8f9fa; border: 1px solid #e9ecef; border-radius: 5px; font-family: 'Courier New', Courier, monospace; font-size: 14px; white-space: pre-wrap; line-height: 1.6; }}
                .log-line {{ margin-bottom: 5px; }}
                .success {{ color: #155724; background-color: #d4edda; border-left: 5px solid #28a745; padding: 5px 10px; }}
                .error {{ color: #721c24; background-color: #f8d7da; border-left: 5px solid #dc3545; padding: 5px 10px; font-weight: bold; }}
                .warning {{ color: #856404; background-color: #fff3cd; border-left: 5px solid #ffc107; padding: 5px 10px; }}
                .info {{ color: #004085; background-color: #cce5ff; border-left: 5px solid #007bff; padding: 5px 10px; font-weight: bold; margin-top: 15px; margin-bottom: 15px;}}
            </style>
        </head>
        <body>
            <div class="container">
                <p class="header">Relatório de Execução Corporativa</p>
                <p><strong>Data de Execução:</strong> {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}</p>
                <div class="status">Status Geral: {status_geral}</div>
                <div class="log-container">
                    {log_html}
                </div>
            </div>
        </body>
        </html>
        """

        try:
            outlook = win32.Dispatch('outlook.application')
            mail = outlook.CreateItem(0)
            mail.To = EMAIL_DESTINO
            mail.Subject = f"Relatório Automático Punch DR90 - {status_geral}"
            mail.HTMLBody = html_body
            mail.Send()
            self.registrar_log("Log enviado via aplicativo Outlook (Formato HTML).")
        except Exception as e:
            self.registrar_log(f"Falha ao enviar e-mail via Outlook Desktop: {e}")

    def get_col_info(self, display_name):
        """Busca informações da coluna usando lógica fuzzy."""
        # 1. Tenta match exato
        col_info = self.schema_lista.get(display_name)
        if col_info:
            return col_info

        # 2. Tenta match normalizado (sem espaços, minúsculo, sem símbolos)
        normalized_target = self.normalize_key(display_name)
        for k, v in self.schema_lista.items():
            if self.normalize_key(k) == normalized_target:
                return v

        return None

    def _simplify_sharepoint_value(self, value):
        """Converte um valor potencialmente complexo do SharePoint em uma string simples."""
        if value is None:
            return ''
        if isinstance(value, dict):
            if 'results' in value and isinstance(value.get('results'), list):
                return '; '.join([self._simplify_sharepoint_value(v) for v in value['results']])
            return value.get('Title', str(value))
        if isinstance(value, list):
            return '; '.join([str(v) for v in value])
        return value

    def tratar_dados(self, raw_data, colunas_desejadas):
        self.registrar_log("Processando dados recebidos (modo inclusivo com mapeamento)...")
        if not raw_data:
            self.registrar_log("AVISO: Nenhum dado bruto para processar.")
            return pd.DataFrame()

        processed_data = []
        all_extra_columns = set()

        METADATA_BLOCKLIST = {
            '__metadata', 'OData__UIVersionString', 'FirstUniqueAncestorSecurableObject',
            'RoleAssignments', 'AttachmentFiles', 'ContentType', 'FieldValuesAsHtml',
            'FieldValuesAsText', 'FieldValuesForEdit', 'File', 'Folder', 'ParentList',
            'Properties', 'Versions', 'odata.editLink', 'odata.etag', 'odata.id',
            'odata.type', 'GUID', 'ServerRedirectedEmbedUri', 'ServerRedirectedEmbedUrl'
        }

        # Extrai os nomes de exibição finais para reordenar o DataFrame mais tarde
        final_display_names = [col[1] if isinstance(col, tuple) else col for col in colunas_desejadas]

        for item in raw_data:
            new_row = {}
            used_internal_names = set()

            # ETAPA 1: Processar colunas desejadas (com ou sem mapeamento)
            for col_config in colunas_desejadas:
                if isinstance(col_config, tuple):
                    source_name, display_name = col_config
                else:
                    source_name, display_name = col_config, col_config

                col_info = self.get_col_info(source_name)
                internal_name, col_type, value = None, 'Text', None

                if col_info:
                    internal_name = col_info['internal_name']
                    col_type = col_info['type']
                    used_internal_names.add(internal_name)
                    used_internal_names.add(f"{internal_name}Id")
                else:  # Fallback para correspondência normalizada se não estiver no schema
                    normalized_target = self.normalize_key(source_name)
                    for k in item.keys():
                        if self.normalize_key(k) == normalized_target:
                            internal_name, col_type = k, 'Text'
                            used_internal_names.add(internal_name)
                            break

                value = item.get(internal_name) if internal_name else item.get(source_name)

                processed_value = ''
                if col_type in ['User', 'Lookup', 'UserMulti', 'LookupMulti']:
                    processed_value = self._simplify_sharepoint_value(value)
                    if not processed_value:
                        val_id = item.get(f"{internal_name}Id") if internal_name else None
                        if val_id: processed_value = f"ID: {self._simplify_sharepoint_value(val_id)}"
                elif value is not None:
                    processed_value = self._simplify_sharepoint_value(value)

                new_row[display_name] = processed_value

            # ETAPA 2: Processar todas as outras colunas não utilizadas
            extra_cols = {}
            for raw_key, raw_value in item.items():
                if raw_key not in used_internal_names and raw_key not in METADATA_BLOCKLIST:
                    simplified_value = self._simplify_sharepoint_value(raw_value)
                    extra_cols[raw_key] = simplified_value
                    all_extra_columns.add(raw_key)

            new_row.update(extra_cols)
            processed_data.append(new_row)

        if all_extra_columns:
            self.registrar_log(
                f"INFO: {len(all_extra_columns)} colunas extras não mapeadas foram adicionadas ao final do relatório.")

        df = pd.DataFrame(processed_data)

        if not df.empty:
            # Reordenar para garantir que as colunas desejadas venham primeiro, na ordem correta
            final_ordered_columns = final_display_names + sorted(list(all_extra_columns))
            existing_cols = [col for col in final_ordered_columns if col in df.columns]
            df = df[existing_cols]

        self.registrar_log(f"DataFrame criado com {df.shape[0]} linhas e {df.shape[1]} colunas.")
        df.columns = df.columns.str.strip().str.replace(r'\s+', ' ', regex=True)

        # ----- Limpeza e Formatação -----
        self.registrar_log("Limpando e formatando o DataFrame final...")
        for col in df.columns:
            df[col] = df[col].astype(str).fillna('')
            df.loc[df[col].str.contains("error|#error", case=False, na=False), col] = ""

            is_date_col = "Date" in col or "Target" in col
            contains_iso_date = df[col].str.contains(r'\d{4}-\d{2}-\d{2}T\d{2}:\d{2}:\d{2}Z', na=False).any()

            if is_date_col or contains_iso_date:
                try:
                    dt_series = pd.to_datetime(df[col], format='%Y-%m-%dT%H:%M:%SZ', errors='coerce', utc=True)
                    valid_dates = dt_series.notna()
                    df.loc[valid_dates, col] = dt_series[valid_dates].dt.strftime('%d/%m/%Y')
                    df[col] = df[col].replace(['NaT', 'nan', 'None', ''], pd.NA)
                except Exception:
                    continue

        self.registrar_log("Limpeza e formatação concluídas.")
        return df

    def obter_schema_lista(self, session, base_url, nome_api):
        self.registrar_log(f"Obtendo schema da lista '{nome_api}'...")
        safe_nome_api = urllib.parse.quote(nome_api)
        endpoint = f"{base_url}/_api/web/lists/getbytitle('{safe_nome_api}')/fields"
        headers = {"Accept": "application/json;odata=verbose"}
        try:
            response = session.get(endpoint, headers=headers)
            if response.status_code == 200:
                fields = response.json()['d']['results']
                self.schema_lista = {f['Title']: {
                    'internal_name': f['InternalName'],
                    'type': f.get('TypeAsString', 'Text'),
                    'static_name': f.get('StaticName', '')
                } for f in fields}
                self.registrar_log(f"Schema obtido. {len(self.schema_lista)} campos mapeados.")
                return True
            else:
                self.registrar_log(
                    f"Falha ao obter schema para '{nome_api}'. Status: {response.status_code}")
                return False
        except Exception as e:
            self.registrar_log(f"Erro ao obter schema para '{nome_api}': {e}")
            return False

    def fetch_sharepoint_data_robustly(self, session, base_url, nome_api, expand_parts, user_fields):
        """
        Tenta baixar dados de forma agressiva (com expands) e, se falhar com 400,
        muda para uma estratégia de enriquecimento em duas etapas.
        """
        safe_nome_api = urllib.parse.quote(nome_api)
        headers = {"Accept": "application/json;odata=verbose"}

        # --- TENTATIVA 1: Consulta Agressiva ---
        self.registrar_log("Tentando download completo com expansão de campos...")
        query_params = ["$top=5000"]
        select_clause = "Id,*"
        expanded_selects = [f"{exp}/Title" for exp in expand_parts]
        if expanded_selects:
            select_clause += "," + ",".join(expanded_selects)
        query_params.append(f"$select={select_clause}")

        if expand_parts:
            expand_str = ','.join(list(set(expand_parts)))
            query_params.append(f"$expand={expand_str}")

        endpoint = f"{base_url}/_api/web/lists/getbytitle('{safe_nome_api}')/items?{'&'.join(query_params)}"

        response = session.get(endpoint, headers=headers)

        if response.status_code == 200:
            self.registrar_log("SUCESSO: Download completo bem-sucedido na primeira tentativa.")
            return response.json().get('d', {}).get('results', [])

        if response.status_code != 400:
            self.registrar_log(f"ERRO: Falha no download com status inesperado: {response.status_code}")
            return None

        # --- TENTATIVA 2: Consulta Base + Enriquecimento ---
        self.registrar_log("AVISO: Erro 400 detectado. Mudando para modo de enriquecimento robusto.")

        # 1. Obter dados base (sem expands)
        endpoint_base = f"{base_url}/_api/web/lists/getbytitle('{safe_nome_api}')/items?$select=Id,*&$top=5000"
        self.registrar_log("Baixando dados base (sem expansão)...")
        response_base = session.get(endpoint_base, headers=headers)

        if response_base.status_code != 200:
            self.registrar_log(f"ERRO CRÍTICO: Falha ao baixar dados base. Status: {response_base.status_code}")
            return None

        base_results = response_base.json().get('d', {}).get('results', [])
        self.registrar_log(f"Dados base com {len(base_results)} itens baixados com sucesso.")

        # 2. Coletar todos os IDs de usuário únicos que precisam ser resolvidos
        user_ids_to_resolve = set()
        for field_internal_name in user_fields:
            id_field = f"{field_internal_name}Id"
            for item in base_results:
                user_id_data = item.get(id_field)
                if user_id_data:
                    # Tratar campos multi-usuário (o ID vem numa lista dentro de um dict)
                    if isinstance(user_id_data, dict) and 'results' in user_id_data:
                        for uid in user_id_data['results']:
                            user_ids_to_resolve.add(uid)
                    # Tratar campos de usuário único
                    else:
                        user_ids_to_resolve.add(user_id_data)

        if not user_ids_to_resolve:
            self.registrar_log("Nenhum ID de usuário para enriquecer. Retornando dados base.")
            return base_results

        self.registrar_log(f"Encontrados {len(user_ids_to_resolve)} IDs de usuário únicos para enriquecer.")

        # 3. Buscar os nomes dos usuários em lotes
        user_id_map = {}
        user_ids_list = list(user_ids_to_resolve)
        batch_size = 100

        for i in range(0, len(user_ids_list), batch_size):
            batch = user_ids_list[i:i + batch_size]
            self.registrar_log(f"Buscando lote de usuários: {i + 1} a {i + len(batch)}...")

            filter_query = " or ".join([f"Id eq {uid}" for uid in batch])
            user_endpoint = f"{base_url}/_api/web/SiteUserInfoList/items?$select=Id,Title&$filter={filter_query}"

            try:
                user_response = session.get(user_endpoint, headers=headers)
                if user_response.status_code == 200:
                    user_data = user_response.json().get('d', {}).get('results', [])
                    for user in user_data:
                        user_id_map[user['Id']] = user['Title']
                else:
                    self.registrar_log(f"AVISO: Falha ao buscar lote de usuários. Status: {user_response.status_code}")
            except Exception as e:
                self.registrar_log(f"ERRO ao buscar lote de usuários: {e}")

        self.registrar_log(f"{len(user_id_map)} de {len(user_ids_to_resolve)} nomes de usuário resolvidos com sucesso.")

        # 4. Enriquecer os resultados base com os nomes
        if not user_id_map:
            self.registrar_log("AVISO: Nenhum nome de usuário pôde ser resolvido. A planilha pode conter apenas IDs.")
            return base_results

        for item in base_results:
            for field_internal_name in user_fields:
                id_field = f"{field_internal_name}Id"
                user_id_data = item.get(id_field)

                if not user_id_data:
                    continue

                # Trata campo multi-usuário
                if isinstance(user_id_data, dict) and 'results' in user_id_data:
                    enriched_users = []
                    for uid in user_id_data['results']:
                        if uid in user_id_map:
                            enriched_users.append({'Title': user_id_map[uid]})
                    # Recria a estrutura que a função tratar_dados espera
                    item[field_internal_name] = {'results': enriched_users}
                # Trata campo de usuário único
                elif user_id_data in user_id_map:
                    item[field_internal_name] = {'Title': user_id_map[user_id_data]}

        self.registrar_log("Enriquecimento de dados concluído.")
        return base_results

    def _sanitize_header(self, header_text):
        """Remove caracteres inválidos para cabeçalhos de tabela do Excel e garante unicidade."""
        if not isinstance(header_text, str):
            header_text = str(header_text)
        # Remove caracteres inválidos para nomes de tabelas/cabeçalhos do Excel
        invalid_chars = r'[\[\]/\\*?:\']'
        sanitized = re.sub(invalid_chars, '', header_text)
        # Trunca para o limite de 255 caracteres do Excel
        return sanitized[:255]

    def iniciar_sessao_navegador(self):
        if not os.path.exists(CAMINHO_DRIVER_FIXO):
            self.registrar_log(f"ERRO CRÍTICO: Driver não encontrado em {CAMINHO_DRIVER_FIXO}")
            return

        edge_options = Options()
        edge_options.add_argument("--ignore-certificate-errors")

        try:
            service = EdgeService(executable_path=CAMINHO_DRIVER_FIXO)
            self.driver = webdriver.Edge(service=service, options=edge_options)
            self.driver.get(URL_LOGIN_SEATRIUM)

            self.registrar_log("Aguardando login na Seatrium...")
            wait = WebDriverWait(self.driver, 120)

            wait.until(EC.presence_of_element_located((
                By.CSS_SELECTOR,
                "[role='grid'], #O365_MainLink_Me, #O365_HeaderLeftRegion, #spCommandBar"
            )))

            self.registrar_log("Sessão autenticada detectada.")
        except Exception as e:
            self.registrar_log(f"Erro no navegador: {e}")

    def extrair_dados(self):
        self.log_sessao = []
        self.registrar_log("Iniciando ciclo de extração...")
        ciclo_sucesso = True

        try:
            if not self.driver:
                self.registrar_log("Driver não inicializado.")
                return

            cookies = self.driver.get_cookies()
            session = requests.Session()
            session.verify = False
            for cookie in cookies:
                session.cookies.set(cookie['name'], cookie['value'])

            # Verifica e cria todas as pastas de destino se não existirem
            for pasta in PASTAS_DESTINO:
                if not os.path.exists(pasta):
                    try:
                        os.makedirs(pasta)
                        self.registrar_log(f"Pasta criada: {pasta}")
                    except Exception as e:
                        self.registrar_log(f"ERRO ao criar pasta {pasta}: {e}")

            for nome_lista, config in LISTAS_SHAREPOINT.items():
                self.registrar_log(f"--- Iniciando processamento da lista: {nome_lista} ---")
                nome_api = config["nome_api"]
                arquivo_saida = config["arquivo_saida"]
                colunas_desejadas = config["colunas"]

                try:
                    safe_nome_api = urllib.parse.quote(nome_api)

                    if not self.obter_schema_lista(session, URL_BASE_SHAREPOINT, nome_api):
                        ciclo_sucesso = False
                        continue

                    self.registrar_log(f"Construindo query para '{nome_api}'...")

                    expand_parts = []
                    user_fields = []  # Lista para os nomes internos dos campos de usuário
                    missing_columns = []

                    # 1. Identificar colunas complexas (User/Lookup) e de usuário
                    for col_config in colunas_desejadas:
                        # Se for uma tupla de mapeamento (source, display), use a origem
                        source_name = col_config[0] if isinstance(col_config, tuple) else col_config
                        col_info = self.get_col_info(source_name)

                        if not col_info:
                            missing_columns.append(nome_coluna)
                            continue

                        internal_name = col_info['internal_name']
                        col_type = col_info['type']

                        if col_type in ['User', 'Lookup', 'UserMulti', 'LookupMulti']:
                            expand_parts.append(internal_name)
                            if col_type in ['User', 'UserMulti']:
                                user_fields.append(internal_name)

                    if missing_columns:
                        self.registrar_log(
                            f"AVISO: {len(missing_columns)} colunas não encontradas no Schema: {', '.join(missing_columns)}")

                    # 2. Chamar a nova função de busca de dados robusta
                    results = self.fetch_sharepoint_data_robustly(
                        session, URL_BASE_SHAREPOINT, nome_api, expand_parts, user_fields
                    )

                    if results is not None:  # A função retorna None em caso de falha crítica
                        if results:
                            df_final = self.tratar_dados(results, colunas_desejadas)

                            # --- LOOP PARA SALVAR EM MÚLTIPLOS DESTINOS ---
                            for pasta_destino in PASTAS_DESTINO:
                                caminho_final = os.path.join(pasta_destino, arquivo_saida)
                                try:
                                    # Cria pasta se não existir na hora H (garantia extra)
                                    if not os.path.exists(pasta_destino):
                                        os.makedirs(pasta_destino)

                                    df_final.to_excel(caminho_final, index=False)
                                    self.registrar_log(f"SUCESSO: Planilha '{nome_lista}' salva em: {caminho_final}")
                                except PermissionError:
                                    self.registrar_log(
                                        f"ERRO DE PERMISSÃO: O arquivo '{arquivo_saida}' está aberto em {pasta_destino}. Feche-o.")
                                    # Não marcamos ciclo_sucesso = False aqui para permitir que salve nas outras pastas se possível
                                    # Mas se for crítico, pode descomentar a linha abaixo:
                                    # ciclo_sucesso = False
                                except Exception as e_save:
                                    self.registrar_log(f"ERRO ao salvar arquivo em {pasta_destino}: {e_save}")
                                    ciclo_sucesso = False
                        else:
                            self.registrar_log(f"AVISO: A lista '{nome_lista}' está vazia.")
                    else:
                        ciclo_sucesso = False

                except Exception as e_lista:
                    self.registrar_log(f"ERRO CRÍTICO na lista '{nome_lista}': {e_lista}")
                    ciclo_sucesso = False
                finally:
                    self.registrar_log(f"--- Fim lista: {nome_lista} ---\n")

            # Após o download de todas as listas, inicia a formatação
            self.formatar_arquivos_como_tabela()

        except Exception as e_ciclo:
            self.registrar_log(f"Falha crítica no ciclo: {e_ciclo}")
            ciclo_sucesso = False
        finally:
            self.enviar_via_outlook_app(ciclo_sucesso)

    def formatar_arquivos_como_tabela(self):
        """
        Percorre pastas, limpa cabeçalhos e formata os dados como Tabela 'Tabela_query'.
        """
        self.registrar_log("--- Iniciando formatação de tabelas (com limpeza de cabeçalho) ---")
        estilo = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False, showLastColumn=False,
                                showRowStripes=True, showColumnStripes=False)

        for pasta in PASTAS_DESTINO:
            if not os.path.exists(pasta):
                self.registrar_log(f"AVISO: Pasta de formatação '{pasta}' não encontrada. Pulando...")
                continue
            self.registrar_log(f"Verificando arquivos para formatação em: {pasta}")

            for config_lista in LISTAS_SHAREPOINT.values():
                arquivo_nome = config_lista["arquivo_saida"]
                caminho_completo = os.path.join(pasta, arquivo_nome)
                if not os.path.exists(caminho_completo):
                    self.registrar_log(f"AVISO: Arquivo '{arquivo_nome}' não encontrado em '{pasta}'.")
                    continue

                try:
                    wb = openpyxl.load_workbook(caminho_completo)
                    sheet = wb.active
                    if sheet.max_row < 1:
                        self.registrar_log(f"AVISO: Arquivo '{arquivo_nome}' está vazio ou sem cabeçalhos.")
                        continue

                    if "Tabela_query" in sheet.tables:
                        self.registrar_log(f"INFO: Arquivo '{arquivo_nome}' já possui 'Tabela_query' formatada.")
                        continue

                    # --- Lógica de Limpeza e Desduplicação de Cabeçalho ---
                    headers = [cell.value for cell in sheet[1]]
                    novos_headers = []
                    seen_headers = set()

                    for header in headers:
                        sanitized = self._sanitize_header(header)

                        # Garante unicidade
                        final_header = sanitized
                        counter = 2
                        while final_header in seen_headers:
                            final_header = f"{sanitized}_{counter}"
                            counter += 1

                        novos_headers.append(final_header)
                        seen_headers.add(final_header)

                    # Escreve os cabeçalhos limpos de volta na planilha
                    for col_idx, new_header_text in enumerate(novos_headers, 1):
                        sheet.cell(row=1, column=col_idx, value=new_header_text)

                    # Se houver tabelas existentes com outros nomes, removemos para evitar conflitos
                    if sheet.tables:
                        for table_name in list(sheet.tables.keys()):
                            self.registrar_log(
                                f"INFO: Removendo tabela antiga '{table_name}' para recriar com cabeçalhos limpos.")
                            del sheet.tables[table_name]

                    # Cria a nova tabela com os cabeçalhos já limpos
                    if sheet.max_row > 0:
                        referencia = f"A1:{get_column_letter(sheet.max_column)}{sheet.max_row}"
                        tab = Table(displayName="Tabela_query", ref=referencia)
                        tab.tableStyleInfo = estilo
                        sheet.add_table(tab)
                        self.registrar_log(f"SUCESSO: Cabeçalhos limpos e 'Tabela_query' criada em '{arquivo_nome}'.")
                    else:
                        self.registrar_log(f"AVISO: Sem dados para criar a tabela em '{arquivo_nome}'.")

                    wb.save(caminho_completo)

                except Exception as e:
                    self.registrar_log(f"ERRO CRÍTICO ao formatar '{arquivo_nome}': {e}")
        self.registrar_log("--- Formatação de tabelas concluída ---")

    def executar(self):
        self.iniciar_sessao_navegador()
        if self.driver:
            self.extrair_dados()
            while True:
                print("Próximo ciclo em 10 minutos...")
                time.sleep(600)
                self.extrair_dados()


if __name__ == "__main__":
    AutomacaoPunchList().executar()
