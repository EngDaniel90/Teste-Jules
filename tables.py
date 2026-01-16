import os
import openpyxl
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter

def formatar_como_tabela(caminhos):
    """
    Percorre pastas, abre arquivos Excel e formata os dados como Tabela.
    Nome da tabela: Tabela_query.
    """
    
    # Estilo da tabela (Azul Médio, padrão do Excel)
    estilo = TableStyleInfo(
        name="TableStyleMedium9", 
        showFirstColumn=False,
        showLastColumn=False, 
        showRowStripes=True, 
        showColumnStripes=False
    )

    for pasta in caminhos:
        if not os.path.exists(pasta):
            print(f"Aviso: A pasta '{pasta}' não foi encontrada. Pulando...")
            continue
            
        print(f"--- Processando pasta: {pasta} ---")
        
        # Lista todos os arquivos na pasta
        for arquivo in os.listdir(pasta):
            if arquivo.lower().endswith(".xlsx"):
                caminho_completo = os.path.join(pasta, arquivo)
                try:
                    wb = openpyxl.load_workbook(caminho_completo)
                    alterou_algo = False
                    
                    # Contador para garantir nomes únicos dentro do MESMO arquivo
                    # (O Excel trava se houver duas tabelas com o mesmo nome no mesmo arquivo)
                    contador_tabelas = 0

                    for sheet in wb.worksheets:
                        # Verifica se a aba tem dados
                        if sheet.max_row < 2 or sheet.max_column < 1:
                            continue # Pula abas vazias ou com apenas cabeçalho sem dados

                        # Define o nome da tabela
                        if contador_tabelas == 0:
                            nome_tabela = "Tabela_query"
                        else:
                            # Se houver mais de uma aba, adiciona sufixo para não corromper o arquivo
                            nome_tabela = f"Tabela_query_{contador_tabelas}"

                        # Verifica se já existe uma tabela para não duplicar
                        if sheet.tables:
                            # Se já tem tabela, apenas tentamos renomear a primeira encontrada
                            tabela_existente = list(sheet.tables.values())[0]
                            tabela_existente.name = nome_tabela
                            tabela_existente.displayName = nome_tabela
                            alterou_algo = True
                            contador_tabelas += 1
                            continue

                        # Calcula o intervalo dos dados (Ex: A1:D50)
                        ultima_linha = sheet.max_row
                        ultima_coluna = sheet.max_column
                        coluna_letra = get_column_letter(ultima_coluna)
                        referencia = f"A1:{coluna_letra}{ultima_linha}"

                        # Cria o objeto Tabela
                        tab = Table(displayName=nome_tabela, ref=referencia)
                        tab.tableStyleInfo = estilo
                        
                        # Adiciona a tabela à aba
                        sheet.add_table(tab)
                        contador_tabelas += 1
                        alterou_algo = True
                        
                    if alterou_algo:
                        wb.save(caminho_completo)
                        print(f"[OK] Formatado: {arquivo}")
                    else:
                        print(f"[--] Nada a alterar: {arquivo}")

                except Exception as e:
                    print(f"[ERRO] Falha ao processar {arquivo}: {e}")

# --- Configuração dos Caminhos ---
# Usamos 'r' antes das aspas para o Python entender as barras invertidas do Windows
pastas_para_processar = [
    r"C:\Users\E797\Downloads\Teste mensagem e print",
    r"C:\Users\E797\PETROBRAS\SRGE SI-II SCP85 ES - Planilha_BI_Punches"
]

if __name__ == "__main__":
    formatar_como_tabela(pastas_para_processar)
    print("\nProcesso finalizado!")
