
import asyncio
import os
import shutil
import flet as ft
from datetime import datetime, date
from sqlalchemy import Column, Integer, String, Boolean, ForeignKey, DateTime, select, update, delete, Table, desc, func
from sqlalchemy.ext.asyncio import create_async_engine, AsyncSession
from sqlalchemy.orm import sessionmaker, declarative_base, relationship, selectinload
from reportlab.lib.pagesizes import A4
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table as RLTable, TableStyle
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors as rl_colors
import openpyxl
from pypdf import PdfWriter

Base = declarative_base()

# Junction Tables
group_participants = Table(
    'group_participants', Base.metadata,
    Column('group_id', Integer, ForeignKey('groups.id'), primary_key=True),
    Column('participant_id', Integer, ForeignKey('participants.id'), primary_key=True)
)

meeting_tasks = Table(
    'meeting_tasks', Base.metadata,
    Column('meeting_id', Integer, ForeignKey('meetings.id'), primary_key=True),
    Column('task_id', Integer, ForeignKey('tasks.id'), primary_key=True)
)

attendance = Table(
    'attendance', Base.metadata,
    Column('meeting_id', Integer, ForeignKey('meetings.id'), primary_key=True),
    Column('participant_id', Integer, ForeignKey('participants.id'), primary_key=True),
    Column('present', Boolean, default=True)
)

class StatusEnum:
    OPEN = "OPEN"
    CLOSED = "CLOSED"

class Setting(Base):
    __tablename__ = 'settings'
    id = Column(Integer, primary_key=True)
    key = Column(String, unique=True)
    value = Column(String)

class Group(Base):
    __tablename__ = 'groups'
    id = Column(Integer, primary_key=True)
    name = Column(String, unique=True, nullable=False)
    description = Column(String)
    participants = relationship("Participant", secondary=group_participants, back_populates="groups")
    meetings = relationship("Meeting", back_populates="group_rel")

class Participant(Base):
    __tablename__ = 'participants'
    id = Column(Integer, primary_key=True)
    name = Column(String, nullable=False)
    email = Column(String)
    company = Column(String)
    groups = relationship("Group", secondary=group_participants, back_populates="participants")
    tasks_assigned = relationship("Task", back_populates="responsible")

class Meeting(Base):
    __tablename__ = 'meetings'
    id = Column(Integer, primary_key=True)
    title = Column(String, nullable=False)
    date = Column(DateTime, default=datetime.now)
    group_id = Column(Integer, ForeignKey('groups.id'))
    attachment_path = Column(String, nullable=True)
    group_rel = relationship("Group", back_populates="meetings")
    tasks_rel = relationship("Task", secondary=meeting_tasks, back_populates="tasks_rel")

class Task(Base):
    __tablename__ = 'tasks'
    id = Column(Integer, primary_key=True)
    description = Column(String, nullable=False)
    status = Column(String, default=StatusEnum.OPEN)
    participant_id = Column(Integer, ForeignKey('participants.id'))
    deadline_1 = Column(DateTime, nullable=True)
    deadline_2 = Column(DateTime, nullable=True)
    deadline_3 = Column(DateTime, nullable=True)
    responsible = relationship("Participant", back_populates="tasks_assigned")
    meetings_rel = relationship("Meeting", secondary=meeting_tasks, back_populates="tasks_rel")

class DBManager:
    def __init__(self, db_url="sqlite+aiosqlite:///atamaster.db"):
        self.engine = create_async_engine(db_url)
        self.async_session = sessionmaker(self.engine, expire_on_commit=False, class_=AsyncSession)

    async def init_db(self):
        async with self.engine.begin() as conn:
            await conn.run_sync(Base.metadata.create_all)

    def to_dict(self, obj):
        if obj is None: return None
        return {c.name: getattr(obj, c.name) for c in obj.__table__.columns}

    async def add_group(self, name, description):
        async with self.async_session() as session:
            g = Group(name=name, description=description)
            session.add(g); await session.commit(); return self.to_dict(g)

    async def get_groups(self):
        async with self.async_session() as session:
            res = await session.execute(select(Group))
            return [self.to_dict(g) for g in res.scalars().all()]

    async def add_participant(self, name, email, company):
        async with self.async_session() as session:
            p = Participant(name=name, email=email, company=company)
            session.add(p); await session.commit(); return self.to_dict(p)

    async def get_participants(self):
        async with self.async_session() as session:
            res = await session.execute(select(Participant))
            return [self.to_dict(p) for p in res.scalars().all()]

    async def get_participant(self, p_id):
        async with self.async_session() as session:
            res = await session.execute(select(Participant).filter(Participant.id == p_id))
            return self.to_dict(res.scalars().first())

    async def add_participant_to_group(self, p_id, g_id):
        async with self.async_session() as session:
            stmt = group_participants.insert().values(group_id=g_id, participant_id=p_id)
            await session.execute(stmt); await session.commit()

    async def get_group_participants(self, g_id):
        async with self.async_session() as session:
            res = await session.execute(select(Participant).join(group_participants).filter(group_participants.c.group_id == g_id))
            return [self.to_dict(p) for p in res.scalars().all()]

    async def get_open_tasks_for_group(self, g_id):
        async with self.async_session() as session:
            res = await session.execute(
                select(Task).join(meeting_tasks).join(Meeting).filter(Meeting.group_id == g_id, Task.status == StatusEnum.OPEN).distinct()
            )
            return [self.to_dict(t) for t in res.scalars().all()]

    async def create_meeting(self, title, group_id, task_data, attendance_data, attachment_path=None):
        async with self.async_session() as session:
            meeting = Meeting(title=title, group_id=group_id, attachment_path=attachment_path)
            session.add(meeting); await session.flush()
            for t_info in task_data:
                if t_info.get("id"):
                    task = await session.get(Task, t_info["id"])
                else:
                    task = Task(
                        description=t_info["description"],
                        participant_id=t_info["participant_id"],
                        deadline_1=t_info.get("deadline_1"),
                        deadline_2=t_info.get("deadline_2"),
                        deadline_3=t_info.get("deadline_3")
                    )
                    session.add(task); await session.flush()
                meeting.tasks_rel.append(task)
            for p_id, present in attendance_data.items():
                await session.execute(attendance.insert().values(meeting_id=meeting.id, participant_id=p_id, present=present))
            await session.commit(); return meeting.id

    async def get_meetings(self, search=""):
        async with self.async_session() as session:
            stmt = select(Meeting).options(selectinload(Meeting.group_rel)).order_by(desc(Meeting.date))
            if search: stmt = stmt.filter(Meeting.title.ilike(f"%{search}%"))
            res = await session.execute(stmt)
            return [{**self.to_dict(m), "group_name": m.group_rel.name if m.group_rel else "N/A"} for m in res.scalars().all()]

    async def get_meeting_details(self, m_id):
        async with self.async_session() as session:
            res = await session.execute(select(Meeting).options(selectinload(Meeting.group_rel), selectinload(Meeting.tasks_rel)).filter(Meeting.id == m_id))
            m = res.scalars().first()
            if not m: return None
            att_res = await session.execute(select(attendance).filter(attendance.c.meeting_id == m_id))
            att = {r.participant_id: r.present for r in att_res.all()}
            tasks = [self.to_dict(t) for t in m.tasks_rel]
            return {**self.to_dict(m), "group_name": m.group_rel.name if m.group_rel else "N/A", "tasks": tasks, "attendance": att}

    async def close_task(self, t_id):
        async with self.async_session() as session:
            await session.execute(update(Task).where(Task.id == t_id).values(status=StatusEnum.CLOSED))
            await session.commit()

    async def update_task_details(self, t_id, description):
        async with self.async_session() as session:
            await session.execute(update(Task).where(Task.id == t_id).values(description=description))
            await session.commit()

    async def get_critical_tasks(self):
        async with self.async_session() as session:
            res = await session.execute(select(Task).filter(Task.status == StatusEnum.OPEN, Task.deadline_3 < datetime.now()))
            return [self.to_dict(t) for t in res.scalars().all()]

    async def get_all_open_tasks_with_info(self):
        async with self.async_session() as session:
            res = await session.execute(
                select(Task, Participant.name, Group.name).join(Participant, Task.participant_id == Participant.id).join(meeting_tasks, Task.id == meeting_tasks.c.task_id).join(Meeting, meeting_tasks.c.meeting_id == Meeting.id).join(Group, Meeting.group_id == Group.id).filter(Task.status == StatusEnum.OPEN).distinct()
            )
            data = []
            for row in res.all():
                d = self.to_dict(row[0]); d['resp_name'] = row[1]; d['group_name'] = row[2]
                data.append(d)
            return data

    async def get_setting(self, key, default=None):
        async with self.async_session() as session:
            res = await session.execute(select(Setting).filter(Setting.key == key))
            s = res.scalars().first()
            return s.value if s else default

    async def update_setting(self, key, value):
        async with self.async_session() as session:
            res = await session.execute(select(Setting).filter(Setting.key == key))
            s = res.scalars().first()
            if s: s.value = value
            else: session.add(Setting(key=key, value=value))
            await session.commit()

# --- UI COMPONENTS ---

class ProfessionalSnackBar(ft.SnackBar):
    def __init__(self, message, success=True):
        super().__init__(content=ft.Text(message, color="white"), bgcolor="green" if success else "#b71c1c", action="OK")

class TaskCard(ft.Container):
    def __init__(self, task, p_name, page):
        super().__init__()
        self.task = task; self.p_name = p_name; self.m_page = page
        self.padding = 15; self.border_radius = 12; self.bgcolor = "#2c2e33"
        status_color = "green" if task['status'] == StatusEnum.CLOSED else "orange"
        is_critical = task['status'] == StatusEnum.OPEN and task['deadline_3'] and task['deadline_3'] < datetime.now()
        if is_critical: self.border = ft.Border.all(2, "red"); status_color = "red"
        d1 = task['deadline_1'].strftime('%d/%m') if task['deadline_1'] else "--"
        d2 = task['deadline_2'].strftime('%d/%m') if task['deadline_2'] else "--"
        d3 = task['deadline_3'].strftime('%d/%m') if task['deadline_3'] else "--"
        self.content = ft.Row([
            ft.Column([ft.Text(task['description'], weight=ft.FontWeight.BOLD, size=16, color="white"), ft.Text(f"Responsável: {p_name}", size=12, color="grey"), ft.Row([self.date_chip(d1, "P1"), self.date_chip(d2, "P2"), self.date_chip(d3, "P3", critical=is_critical)])], expand=True),
            ft.Container(content=ft.Text(task['status'], size=10, weight=ft.FontWeight.BOLD, color="black"), bgcolor=status_color, padding=ft.Padding.symmetric(6, 12), border_radius=15)
        ])
    def date_chip(self, text, label, critical=False):
        color = "#b71c1c" if critical and label == "P3" else "#424242"
        return ft.Container(content=ft.Text(f"{label}: {text}", size=10, color="white"), bgcolor=color, padding=ft.Padding.symmetric(2, 6), border_radius=4)

class DashboardView(ft.Column):
    def __init__(self, page):
        super().__init__(expand=True, scroll=ft.ScrollMode.AUTO, spacing=20); self.m_page = page
    async def refresh(self):
        if not self.page: return
        try:
            critical = await self.m_page.db.get_critical_tasks()
            open_all = await self.m_page.db.get_all_open_tasks_with_info()
            theme_color = self.m_page.theme_color
            self.controls = [
                ft.Column([ft.Text("Dashboard Executivo", size=32, weight=ft.FontWeight.BOLD, color="white"), ft.Text("Gestão de pendências e visão sistêmica", color="grey")]),
                ft.Row([self.stat_card("ITENS CRÍTICOS", str(len(critical)), "red", ft.Icons.WARNING_AMBER), self.stat_card("TOTAL PENDENTE", str(len(open_all)), theme_color, ft.Icons.LIST_ALT)], spacing=20),
                ft.Divider(height=20, color="transparent"),
                ft.Row([ft.Text("Pendências de Alta Prioridade", size=22, weight=ft.FontWeight.BOLD, color="white", expand=True), ft.FilledButton("Relatório Global de Gestão", icon=ft.Icons.FILE_DOWNLOAD, on_click=lambda e: self.m_page.run_task(self.export_excel_report))]),
            ]
            if not critical: self.controls.append(ft.Container(content=ft.Text("Excelente! Nenhuma pendência crítica hoje.", color="grey"), padding=20, border_radius=10, border=ft.Border.all(1, "#333333")))
            for t in critical:
                p = await self.m_page.db.get_participant(t['participant_id'])
                self.controls.append(TaskCard(t, p['name'] if p else "N/A", self.m_page))
            self.update()
        except Exception as ex: self.m_page.snack_bar = ProfessionalSnackBar(f"Erro: {ex}", False); self.m_page.snack_bar.open = True; self.m_page.update()
    async def export_excel_report(self, e=None):
        try:
            tasks = await self.m_page.db.get_all_open_tasks_with_info()
            wb = openpyxl.Workbook(); ws = wb.active; ws.title = "AtaMaster Tasks"
            ws.append(["ID", "Grupo", "Descrição", "Responsável", "Prazo 1", "Prazo 2", "Prazo 3", "Status"])
            for t in tasks: ws.append([t['id'], t['group_name'], t['description'], t['resp_name'], t['deadline_1'].strftime('%d/%m/%Y') if t['deadline_1'] else "", t['deadline_2'].strftime('%d/%m/%Y') if t['deadline_2'] else "", t['deadline_3'].strftime('%d/%m/%Y') if t['deadline_3'] else "", t['status']])
            fname = f"Gestao_Pendencias_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"; wb.save(fname); os.startfile(fname) if os.name == 'nt' else None
            self.m_page.snack_bar = ProfessionalSnackBar("Excel exportado com sucesso!"); self.m_page.snack_bar.open = True; self.m_page.update()
        except Exception as ex: self.m_page.snack_bar = ProfessionalSnackBar(f"Erro: {ex}", False); self.m_page.snack_bar.open = True; self.m_page.update()
    def stat_card(self, title, value, color, icon):
        return ft.Container(content=ft.Column([ft.Row([ft.Icon(icon, color=color, size=20), ft.Text(title, size=12, color="grey", weight=ft.FontWeight.BOLD)]), ft.Text(value, size=40, weight=ft.FontWeight.BOLD, color="white")], spacing=5), bgcolor="#1e1e1e", padding=25, border_radius=15, expand=True, border=ft.Border.all(1, "#333333"))

class ManagementView(ft.Column):
    def __init__(self, page):
        super().__init__(expand=True, scroll=ft.ScrollMode.AUTO); self.m_page = page; self.selected_tab = "groups"
    async def refresh(self, initial=False):
        if not self.page and not initial: return
        self.controls.clear()
        tabs = ft.Container(content=ft.Row([self.tab_btn("Grupos", "groups"), self.tab_btn("Participantes", "participants"), self.tab_btn("Config", "config"), self.tab_btn("Ajuda", "help")], spacing=10), padding=ft.Padding.only(bottom=20))
        content = await self.get_tab_content()
        self.controls = [ft.Text("Administração do Sistema", size=32, weight=ft.FontWeight.BOLD, color="white"), tabs, ft.Container(content=content, padding=10)]
        if not initial: self.update()
    def tab_btn(self, text, key):
        active = self.selected_tab == key
        theme_color = self.m_page.theme_color
        return ft.TextButton(text, on_click=lambda e: self.m_page.run_task(self.set_tab, key), style=ft.ButtonStyle(color=theme_color if active else "white", bgcolor="#333333" if active else "transparent", shape=ft.RoundedRectangleBorder(radius=8)))
    async def set_tab(self, tab): self.selected_tab = tab; await self.refresh()
    async def get_tab_content(self):
        if self.selected_tab == "groups": return await self.group_tab()
        if self.selected_tab == "participants": return await self.participant_tab()
        if self.selected_tab == "config": return await self.config_tab()
        return self.help_tab()
    async def group_tab(self):
        groups = await self.m_page.db.get_groups()
        theme_color = self.m_page.theme_color
        g_list = ft.Column([ft.ListTile(title=ft.Text(g['name'], weight=ft.FontWeight.BOLD, color="white"), subtitle=ft.Text(g['description'], color="grey"), leading=ft.Icon(ft.Icons.GROUP_WORK, color=theme_color)) for g in groups])
        name_i = ft.TextField(label="Nome do Grupo", expand=True); desc_i = ft.TextField(label="Descrição", expand=True)
        async def add_g(e):
            if name_i.value:
                await self.m_page.db.add_group(name_i.value, desc_i.value); await self.refresh()
                self.m_page.snack_bar = ProfessionalSnackBar("Grupo registrado!"); self.m_page.snack_bar.open = True; self.m_page.update()
        return ft.Column([ft.Row([name_i, desc_i, ft.FilledButton("Adicionar", icon=ft.Icons.ADD, on_click=lambda e: self.m_page.run_task(add_g, e))]), ft.Divider(height=40, color="#333333"), g_list])
    async def participant_tab(self):
        ps = await self.m_page.db.get_participants(); gs = await self.m_page.db.get_groups()
        p_list = ft.Column([ft.ListTile(title=ft.Text(p['name'], weight=ft.FontWeight.BOLD, color="white"), subtitle=ft.Text(f"{p['company']} • {p['email']}", color="grey"), leading=ft.Icon(ft.Icons.PERSON, color="indigo")) for p in ps])
        n_i = ft.TextField(label="Nome", expand=True); e_i = ft.TextField(label="Email", expand=True); c_i = ft.TextField(label="Empresa", expand=True)
        g_d = ft.Dropdown(label="Grupo", options=[ft.dropdown.Option(key=str(g['id']), text=g['name']) for g in gs], expand=True)
        async def add_p(e):
            if n_i.value:
                np = await self.m_page.db.add_participant(n_i.value, e_i.value, c_i.value)
                if g_d.value: await self.m_page.db.add_participant_to_group(np['id'], int(g_d.value))
                await self.refresh()
        async def import_excel(e):
            res = await self.m_page.excel_picker.pick_files(allowed_extensions=["xlsx"])
            if res and res.files:
                wb = openpyxl.load_workbook(res.files[0].path); ws = wb.active
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if row[0]: await self.m_page.db.add_participant(str(row[0]), str(row[1]) if len(row)>1 else "", str(row[2]) if len(row)>2 else "")
                await self.refresh()
        return ft.Column([ft.Row([n_i, e_i, c_i]), ft.Row([g_d, ft.FilledButton("Salvar", on_click=lambda e: self.m_page.run_task(add_p, e)), ft.FilledButton("Importar Excel", icon=ft.Icons.UPLOAD_FILE, on_click=lambda e: self.m_page.run_task(import_excel, e))]), ft.Divider(height=30, color="#333333"), p_list])
    async def config_tab(self):
        curr_color = await self.m_page.db.get_setting("theme_color", "cyan")
        color_d = ft.Dropdown(label="Cor de Destaque do Sistema", value=curr_color, options=[ft.dropdown.Option("cyan", "Cyan Corporate"), ft.dropdown.Option("indigo", "Indigo Night"), ft.dropdown.Option("green", "Green Growth")], expand=True)
        async def save_config(e):
            await self.m_page.db.update_setting("theme_color", color_d.value)
            self.m_page.theme_color = color_d.value
            self.m_page.theme.color_scheme_seed = color_d.value
            self.m_page.snack_bar = ProfessionalSnackBar("Configurações salvas!"); self.m_page.snack_bar.open = True; self.m_page.update()

        backup_card = ft.Container(content=ft.Column([ft.Text("Manutenção Preventiva", size=18, weight=ft.FontWeight.BOLD, color="white"), ft.Text("Crie cópias de segurança para garantir a integridade dos dados.", color="grey"), ft.Row([ft.FilledButton("Exportar Banco", icon=ft.Icons.SAVE, on_click=lambda e: self.m_page.run_task(self.m_page.run_backup, e)), ft.FilledButton("Restaurar Backup", icon=ft.Icons.RESTORE, on_click=lambda e: self.m_page.run_task(self.m_page.run_restore, e))], spacing=20)]), padding=20, bgcolor="#1a1a1a", border_radius=10)

        return ft.Column([ft.Row([color_d, ft.FilledButton("Salvar", on_click=lambda e: self.m_page.run_task(save_config, e))]), ft.Divider(height=30), backup_card])

    def help_tab(self):
        return ft.Container(content=ft.Column([
            ft.Text("Manual do Usuário - AtaMaster Pro", size=20, weight=ft.FontWeight.BOLD, color="white"),
            ft.Text("1. Cadastro: Comece criando Grupos e Participantes na aba Gestão.", color="white"),
            ft.Text("2. Nova Ata: Clique em 'NOVA REUNIÃO'. Selecione o Grupo para carregar as tarefas em aberto.", color="white"),
            ft.Text("3. Ata Viva: Itens não concluídos em reuniões anteriores são trazidos automaticamente.", color="white"),
            ft.Text("4. Prazos: O sistema alerta em vermelho quando o terceiro prazo (P3) é atingido.", color="white"),
            ft.Text("5. Documentos: Você pode anexar outros PDFs que serão fundidos à ata final.", color="white"),
            ft.Divider(),
            ft.Text("Suporte: Daniel Alves Anversi", italic=True, color="grey")
        ], spacing=10, scroll=ft.ScrollMode.AUTO), padding=20)

class NewMeetingView(ft.Column):
    def __init__(self, page):
        super().__init__(expand=True, scroll=ft.ScrollMode.AUTO, spacing=15); self.m_page = page
        self.temp_tasks = []; self.attachments = []; self.deadlines = [None, None, None]
        theme_color = self.m_page.theme_color
        self.title_i = ft.TextField(label="Assunto da Reunião", value=f"Reunião de Alinhamento {datetime.now().strftime('%d/%m/%Y')}", border_color=theme_color)
        self.group_d = ft.Dropdown(label="Unidade / Contexto", on_change=lambda e: self.m_page.run_task(self.on_group_select, e), border_color=theme_color)
        self.attendance_col = ft.ResponsiveRow()
        self.task_desc = ft.TextField(label="Descrição do Item", expand=True); self.task_resp = ft.Dropdown(label="Responsável", expand=True)
        self.d1_btn = ft.OutlinedButton("Prazo 1", on_click=lambda _: self.m_page.run_task(self.open_dp, 0))
        self.d2_btn = ft.OutlinedButton("Prazo 2", on_click=lambda _: self.m_page.run_task(self.open_dp, 1))
        self.d3_btn = ft.OutlinedButton("Prazo 3", on_click=lambda _: self.m_page.run_task(self.open_dp, 2))
        self.tasks_list_display = ft.Column(); self.attachment_display = ft.Row(wrap=True)
    async def open_dp(self, idx): self.m_page.active_dp_idx = idx; await self.m_page.date_picker.pick_date()
    async def handle_date_change(self, e):
        idx = self.m_page.active_dp_idx; d = self.m_page.date_picker.value
        if d: self.deadlines[idx] = d; btns = [self.d1_btn, self.d2_btn, self.d3_btn]; btns[idx].text = f"P{idx+1}: {d.strftime('%d/%m/%y')}"; self.update()
    async def refresh(self):
        if not self.page: return
        gs = await self.m_page.db.get_groups(); self.group_d.options = [ft.dropdown.Option(key=str(g['id']), text=g['name']) for g in gs]
        self.controls = [ft.Text("Nova Ata Viva", size=32, weight=ft.FontWeight.BOLD, color="white"), ft.Row([self.title_i, self.group_d], spacing=20), ft.Text("Presença e Ciência", size=18, weight=ft.FontWeight.BOLD, color="white"), self.attendance_col, ft.Text("Adicionar Nova Pendência", size=18, weight=ft.FontWeight.BOLD, color="white"), ft.Row([self.task_desc, self.task_resp]), ft.Row([self.d1_btn, self.d2_btn, self.d3_btn, ft.FilledButton("Adicionar", icon=ft.Icons.ADD, on_click=lambda e: self.m_page.run_task(self.add_task_to_meeting, e))]), ft.Text("Pauta / Histórico Aberto", size=18, weight=ft.FontWeight.BOLD, color="white"), self.tasks_list_display, ft.Text("Fusão de Documentos (Anexar PDF/PPT)", size=18, weight=ft.FontWeight.BOLD, color="white"), ft.FilledButton("Selecionar Anexos", icon=ft.Icons.ATTACH_FILE, on_click=lambda e: self.m_page.run_task(self.pick_attachments, e)), self.attachment_display, ft.Divider(height=20), ft.FilledButton("FINALIZAR ATA E GERAR PDF COMPLETO", icon=ft.Icons.PICTURE_AS_PDF, on_click=lambda e: self.m_page.run_task(self.save_meeting, e), height=55, expand=True)]
        self.update()
    async def pick_attachments(self, e):
        res = await self.m_page.attach_picker.pick_files(allow_multiple=True, allowed_extensions=["pdf"])
        if res and res.files:
            for f in res.files: self.attachments.append(f.path); self.attachment_display.controls.append(ft.Chip(label=ft.Text(os.path.basename(f.path)), on_delete=lambda _, p=f.path: self.remove_attach(p)))
            self.update()
    def remove_attach(self, path): self.attachments.remove(path); self.attachment_display.controls = [c for c in self.attachment_display.controls if getattr(c.label, "value", "") == os.path.basename(path)]; self.update()
    async def on_group_select(self, e):
        if not self.group_d.value: return
        g_id = int(self.group_d.value); gps = await self.m_page.db.get_group_participants(g_id)
        self.attendance_col.controls = [ft.Container(ft.Checkbox(label=p['name'], value=True, data=p['id']), col={"sm": 6, "md": 4, "lg": 3}) for p in gps]
        self.task_resp.options = [ft.dropdown.Option(key=str(p['id']), text=p['name']) for p in gps]
        ots = await self.m_page.db.get_open_tasks_for_group(g_id); self.temp_tasks = [{**ot, "from_db": True} for ot in ots]; await self.refresh_tasks()
    async def add_task_to_meeting(self, e):
        if not self.task_desc.value or not self.task_resp.value: self.m_page.snack_bar = ProfessionalSnackBar("Preencha todos os campos!", False); self.m_page.snack_bar.open = True; self.m_page.update(); return
        self.temp_tasks.append({"description": self.task_desc.value, "participant_id": int(self.task_resp.value), "deadline_1": self.deadlines[0], "deadline_2": self.deadlines[1], "deadline_3": self.deadlines[2], "from_db": False})
        self.task_desc.value = ""; self.deadlines = [None, None, None]; self.d1_btn.text = "Prazo 1"; self.d2_btn.text = "Prazo 2"; self.d3_btn.text = "Prazo 3"; await self.refresh_tasks()
    async def refresh_tasks(self):
        self.tasks_list_display.controls.clear()
        theme_color = self.m_page.theme_color
        for i, t in enumerate(self.temp_tasks):
            p = await self.m_page.db.get_participant(t['participant_id']); p_name = p['name'] if p else "N/A"
            self.tasks_list_display.controls.append(ft.Container(content=ft.Row([ft.Icon(ft.Icons.HISTORY if t.get("from_db") else ft.Icons.ADD_CIRCLE, color=theme_color if t.get("from_db") else "white"), ft.Text(t["description"], expand=True, color="white"), ft.Text(p_name, width=150, color="grey", size=12), ft.IconButton(ft.Icons.DELETE_OUTLINE, icon_color="red", on_click=lambda _, idx=i: self.m_page.run_task(self.remove_task, idx))]), padding=10, bgcolor="#262626", border_radius=8))
        self.update()
    async def remove_task(self, idx): self.temp_tasks.pop(idx); await self.refresh_tasks()
    async def save_meeting(self, e):
        if not self.group_d.value: return
        att_data = {c.content.data: c.content.value for c in self.attendance_col.controls}
        ata_temp = f"temp_ata_{int(datetime.now().timestamp())}.pdf"; await self.generate_ata_pdf(ata_temp, self.title_i.value, int(self.group_d.value), self.temp_tasks, att_data)
        final_fn = f"Ata_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
        try:
            merger = PdfWriter(); merger.append(ata_temp)
            for pdf in self.attachments:
                if os.path.exists(pdf): merger.append(pdf)
            with open(final_fn, "wb") as f: merger.write(f)
            if os.path.exists(ata_temp): os.remove(ata_temp)
        except Exception as ex: final_fn = ata_temp
        await self.m_page.db.create_meeting(self.title_i.value, int(self.group_d.value), self.temp_tasks, att_data, attachment_path=final_fn)
        os.startfile(final_fn) if os.name == 'nt' else None; await self.m_page.push_route("/meetings")
    async def generate_ata_pdf(self, filename, title, group_id, tasks, att_data):
        doc = SimpleDocTemplate(filename, pagesize=A4); styles = getSampleStyleSheet()
        gs = await self.m_page.db.get_groups(); g_name = next((g['name'] for g in gs if g['id'] == group_id), "N/A")
        elements = [Paragraph(f"ATA DE REUNIÃO: {title}", styles['Title']), Paragraph(f"Data: {datetime.now().strftime('%d/%m/%Y')} | Grupo: {g_name}", styles['Normal']), Spacer(1, 24)]
        elements.append(Paragraph("LISTA DE PRESENÇA", styles['Heading2']))
        att_rows = [["Nome", "Presença"]]; [att_rows.append([(await self.m_page.db.get_participant(p_id))['name'], "Sim" if present else "Não"]) for p_id, present in att_data.items()]
        elements.append(RLTable(att_rows, style=TableStyle([('BACKGROUND',(0,0),(-1,0),rl_colors.cyan),('TEXTCOLOR',(0,0),(-1,0),rl_colors.whitesmoke)]))); elements.append(Spacer(1, 24))
        elements.append(Paragraph("ACOMPANHAMENTO DE TAREFAS", styles['Heading2']))
        task_rows = [["Descrição", "Responsável", "Prazo 3", "Status"]]; [task_rows.append([t['description'], (await self.m_page.db.get_participant(t['participant_id']))['name'], t['deadline_3'].strftime('%d/%m/%Y') if t['deadline_3'] else "--", t.get('status', 'OPEN')]) for t in tasks]
        elements.append(RLTable(task_rows, style=TableStyle([('GRID', (0,0), (-1,-1), 0.5, rl_colors.grey)]))); elements.append(Spacer(1, 48))
        elements.append(Paragraph("ASSINATURAS:", styles['Heading2']))
        for p_id, present in att_data.items():
            if present: p = await self.m_page.db.get_participant(p_id); elements.append(Spacer(1, 20)); elements.append(Paragraph("________________________________________________", styles['Normal'])); elements.append(Paragraph(f"{p['name']} ({p['company']})", styles['Normal']))
        doc.build(elements)

class MeetingsView(ft.Column):
    def __init__(self, page):
        super().__init__(expand=True, scroll=ft.ScrollMode.AUTO, spacing=10); self.m_page = page
        self.search_field = ft.TextField(label="Pesquisar atas...", expand=True, on_change=lambda e: self.m_page.run_task(self.refresh))
    async def refresh(self, e=None):
        if not self.page: return
        ms = await self.m_page.db.get_meetings(search=self.search_field.value)
        self.controls = [ft.Text("Histórico Executivo", size=32, weight=ft.FontWeight.BOLD, color="white"), ft.Row([self.search_field])]
        theme_color = self.m_page.theme_color
        for m in ms:
            self.controls.append(ft.Container(content=ft.ListTile(title=ft.Text(m['title'], weight=ft.FontWeight.BOLD, color="white"), subtitle=ft.Text(f"{m['date'].strftime('%d/%m/%Y')} • {m['group_name']}", color="grey"), trailing=ft.Icon(ft.Icons.CHEVRON_RIGHT, color=theme_color), on_click=lambda _, mid=m['id']: self.m_page.run_task(self.m_page.push_route, f"/meeting/{mid}")), bgcolor="#262626", border_radius=10))
        self.update()

class MeetingDetailView(ft.Column):
    def __init__(self, page, m_id):
        super().__init__(expand=True, scroll=ft.ScrollMode.AUTO, spacing=20); self.m_page = page; self.m_id = m_id
    async def refresh(self):
        if not self.page: return
        m = await self.m_page.db.get_meeting_details(self.m_id)
        if not m: return
        self.controls = [ft.Row([ft.IconButton(ft.Icons.ARROW_BACK, on_click=lambda _: self.m_page.run_task(self.m_page.push_route, "/meetings")), ft.Text(m['title'], size=28, weight=ft.FontWeight.BOLD, color="white")]), ft.Text(f"{m['date'].strftime('%d/%m/%Y')} | {m['group_name']}", color="grey"), ft.Row([ft.FilledButton("Ver Documento", icon=ft.Icons.PICTURE_AS_PDF, on_click=lambda _: os.startfile(m['attachment_path']) if m['attachment_path'] and os.name=='nt' else None)]), ft.Divider(color="#333333"), ft.Text("Itens de Acompanhamento", size=20, weight=ft.FontWeight.BOLD, color="white")]
        theme_color = self.m_page.theme_color
        for t in m['tasks']:
            p = await self.m_page.db.get_participant(t['participant_id']); actions = ft.Row()
            if t['status'] == StatusEnum.OPEN: actions.controls.extend([ft.IconButton(ft.Icons.EDIT, icon_color=theme_color, on_click=lambda _, tid=t['id'], desc=t['description']: self.show_edit_dialog(tid, desc)), ft.FilledButton("Concluir", on_click=lambda _, tid=t['id']: self.m_page.run_task(self.close_item, tid), bgcolor="green")])
            self.controls.append(ft.Container(content=ft.Row([ft.Column([ft.Text(t['description'], weight=ft.FontWeight.BOLD, color="white"), ft.Text(f"Responsável: {p['name'] if p else 'N/A'}", size=12, color="grey")], expand=True), ft.Text(t['status'], color=theme_color if t['status']==StatusEnum.OPEN else "green", weight=ft.FontWeight.BOLD), actions]), padding=15, bgcolor="#262626", border_radius=12))
        self.update()
    def show_edit_dialog(self, t_id, old_desc):
        ei = ft.TextField(value=old_desc, expand=True)
        def save_e(e): self.m_page.run_task(self.save_task_edit, t_id, ei.value); self.m_page.dialog.open = False; self.m_page.update()
        self.m_page.dialog = ft.AlertDialog(title=ft.Text("Editar Descrição"), content=ei, actions=[ft.TextButton("Cancelar", on_click=lambda _: setattr(self.m_page.dialog, "open", False)), ft.TextButton("Salvar", on_click=save_e)])
        self.m_page.dialog.open = True; self.m_page.update()
    async def save_task_edit(self, t_id, nd): await self.m_page.db.update_task_details(t_id, nd); await self.refresh()
    async def close_item(self, t_id): await self.m_page.db.close_task(t_id); await self.refresh()

class Sidebar(ft.Container):
    def __init__(self, page):
        super().__init__(); self.m_page = page; self.width = 280; self.bgcolor = "#121212"; self.padding = 20
        self.nav_col = ft.Column(spacing=5); self.content = ft.Column([ft.Container(content=ft.Row([ft.Icon(ft.Icons.DOMAIN, color=self.m_page.theme_color, size=30), ft.Text("ATAMASTER", size=24, weight=ft.FontWeight.BOLD, color="white")]), margin=ft.margin.only(bottom=40, top=10)), self.nav_col, ft.Divider(height=40, color="#333333"), ft.Container(content=ft.Row([ft.Icon(ft.Icons.ADD_CIRCLE, color=self.m_page.theme_color), ft.Text("NOVA REUNIÃO", color=self.m_page.theme_color, weight=ft.FontWeight.BOLD)]), padding=15, border=ft.Border.all(1, self.m_page.theme_color), border_radius=12, on_click=lambda _: self.m_page.run_task(self.m_page.push_route, "/new_meeting"), ink=True), ft.Container(expand=True), ft.Text("Business Intelligence 2.5", size=10, color="#444444"), ft.Text("Daniel Alves Anversi", size=11, italic=True, color="grey")], spacing=10); self.update_nav()
    def update_nav(self): self.nav_col.controls = [self.nav_item(ft.Icons.DASHBOARD_ROUNDED, "Dashboard", "/"), self.nav_item(ft.Icons.DESCRIPTION_ROUNDED, "Histórico", "/meetings"), self.nav_item(ft.Icons.SETTINGS_SUGGEST_ROUNDED, "Gestão", "/management")]
    def nav_item(self, icon, text, route):
        is_active = self.m_page.route == route
        theme_color = self.m_page.theme_color
        return ft.Container(content=ft.Row([ft.Icon(icon, color=theme_color if is_active else "grey"), ft.Text(text, color="white" if is_active else "grey", weight=ft.FontWeight.BOLD if is_active else ft.FontWeight.NORMAL)]), padding=12, border_radius=10, on_click=lambda _, r=route: self.m_page.run_task(self.m_page.push_route, r), ink=True, bgcolor="#222222" if is_active else "transparent")

async def main(page: ft.Page):
    def run_task(coro, *args, **kwargs):
        return asyncio.create_task(coro(*args, **kwargs))
    page.run_task = run_task
    async def push_route(route):
        page.go(route)
    page.push_route = push_route
    page.db = DBManager(); await page.db.init_db()
    page.theme_color = await page.db.get_setting("theme_color", "cyan")
    page.title = "AtaMaster Pro"; page.theme_mode = ft.ThemeMode.DARK; page.padding = 0; page.theme = ft.Theme(font_family="Segoe UI", color_scheme_seed=page.theme_color)
    page.excel_picker = ft.FilePicker(); page.backup_picker = ft.FilePicker(); page.attach_picker = ft.FilePicker(); page.date_picker = ft.DatePicker(on_change=lambda e: page.run_task(page.current_view.handle_date_change, e) if hasattr(page.current_view, 'handle_date_change') else None)
    page.overlay.extend([page.excel_picker, page.backup_picker, page.attach_picker, page.date_picker])
    async def run_backup(e=None):
        path = await page.backup_picker.save_file(file_name="atamaster_backup.db")
        if path: shutil.copy("atamaster.db", path)
    page.run_backup = run_backup
    async def run_restore(e=None):
        res = await page.backup_picker.pick_files(allowed_extensions=["db"])
        if res and res.files: shutil.copy(res.files[0].path, "atamaster.db"); await page.db.init_db(); await page.push_route("/")
    page.run_restore = run_restore
    content_container = ft.Container(expand=True, padding=40, bgcolor="#0f0f0f"); sidebar = Sidebar(page)
    async def route_change(e):
        r = page.route
        if r == "/": v = DashboardView(page)
        elif r == "/meetings": v = MeetingsView(page)
        elif r == "/management": v = ManagementView(page)
        elif r == "/new_meeting": v = NewMeetingView(page)
        elif r.startswith("/meeting/"): v = MeetingDetailView(page, int(r.split("/")[-1]))
        else: v = ft.Text("404")
        page.current_view = v; content_container.content = v; sidebar.update_nav(); page.update()
        if hasattr(v, 'refresh'): await v.refresh()
        page.update()
    page.on_route_change = route_change; layout = ft.Row([sidebar, ft.VerticalDivider(width=1, color="#222222"), content_container], expand=True, spacing=0); page.add(layout); await page.push_route("/")

if __name__ == "__main__":
    ft.run(main)
